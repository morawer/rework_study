import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import os
from datetime import datetime


def graphsAvgCreator():

    weekNumberAux = 0
    linesAcum = 0
    linesAvg = 0
    linesRow = 0
    dates = []
    lines = []
    counterAHU = []
    weeksAcum = 0
    initWeek = True

    path = '/home/dani/projects/rework_study/dataAHU.xlsx'

    if os.path.exists(path):
        wb = load_workbook(path)
        sheet = wb.active
        lastRow = sheet.max_row

        for row in range(2, lastRow + 1):
            cell_obj = sheet.cell(row=row, column=1)
            date = cell_obj.value
            dateformat = datetime.strptime(date, '%d/%m/%Y')
            weeknumber = dateformat.strftime('%W')

            if (weeknumber != weekNumberAux or initWeek == True or row == lastRow):
                if row > 2 and row != lastRow:
                    dates.append(weekNumberAux)
                    linesAvg = int(linesAcum)/int(weeksAcum)
                    lines.append(float("{:.1f}".format(linesAvg)))
                    counterAHU.append(weeksAcum)
                    linesAvg = 0
                    linesAcum = 0
                    weeksAcum = 0
                    initWeek = True

                elif row == lastRow:
                    weekNumberAux = weeknumber
                    linesRow = sheet.cell(row=row, column=6)
                    linesInt = linesRow.value
                    linesAcum = int(linesAcum) + int(linesInt)
                    weeksAcum = weeksAcum + 1
                    initWeek = False
                    dates.append(weeknumber)
                    linesAvg = int(linesAcum)/int(weeksAcum)
                    lines.append(float("{:.1f}".format(linesAvg)))
                    counterAHU.append(weeksAcum)
                    break

                weekNumberAux = weeknumber
                linesAcum = 0
                linesRow = sheet.cell(row=row, column=6)
                linesInt = linesRow.value
                linesAcum = int(linesAcum) + int(linesInt)
                weeksAcum = weeksAcum + 1
                initWeek = False
            else:
                linesRow = sheet.cell(row=row, column=6)
                linesInt = linesRow.value
                linesAcum = int(linesAcum) + int(linesInt)
                weeksAcum = weeksAcum + 1
                initWeek = False
            if len(dates) == 11:
                dates.remove(dates[0])
                lines.remove(lines[0])
                counterAHU.remove(counterAHU[0])
    else:
        print('El archivo no existe')

    def totalAvgLines():
        avgAcum = 0
        for avg in lines:
            avgAcum = avgAcum + avg
        return avgAcum / len(lines)

    totalAvgLines = totalAvgLines()
    fig, ax = plt.subplots()
    x = np.arange(len(dates))
    width = 0.35
    bar1 = ax.bar(x - width/2, counterAHU, width, label="Nº UTA's")
    bar2 = ax.bar(x + width/2, lines, width, label="Media de lineas")

    ax.set_xlabel('Semanas')
    ax.set_title(
        f"Número de UTA's y media de líneas por semana | totalAvg: {totalAvgLines:.1f}")
    ax.set_xticks(x)
    ax.set_xticklabels(dates)

    ax.axhline(y=totalAvgLines, linewidth=3, color='r')
    ax.legend()

    def autolabel(bars):
        for bar in bars:
            height = bar.get_height()
            ax.annotate('{}'.format(height),
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom')

    autolabel(bar1)
    autolabel(bar2)
    fig.tight_layout()

    dateGraph = datetime.now()
    dateGraphWeekNumber = int(dateGraph.strftime('%W')) - 1
    dateGraphWeek = int(dateGraphWeekNumber)
    plt.savefig(
        f'/home/dani/projects/rework_study/avg_week_{dateGraphWeek}_graph')
