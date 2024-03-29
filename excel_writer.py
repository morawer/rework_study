from fileinput import filename
from importlib.resources import path
from turtle import mode
from openpyxl import load_workbook
from openpyxl import Workbook
import os

def excelWriter(dataArray):
    
    fileName = '/home/dani/projects/rework_study/dataAHU.xlsx'
    
    if os.path.exists(fileName):
        wb = load_workbook(fileName)
    else:
        wb = Workbook()
    
    sheet = wb.active
    sheet.title = 'dataAHU'
    lastRow = sheet.max_row
    lenghtData = len(dataArray.tags)

    sheet.cell(row=lastRow+1, column=1, value=dataArray.date)
    sheet.cell(row=lastRow+1, column=2, value=dataArray.order)
    sheet.cell(row=lastRow+1, column=3, value=dataArray.mo)
    sheet.cell(row=lastRow+1, column=4, value=dataArray.model)
    sheet.cell(row=lastRow+1, column=5, value=dataArray.inspector)
    sheet.cell(row=lastRow+1, column=6, value=dataArray.lines)
    
    for data in range(0, lenghtData):
        sheet.cell(row=lastRow+1, column=7+data, value=dataArray.tags[data])
        
    wb.save(fileName)
    wb.close()
